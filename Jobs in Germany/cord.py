# SAVE DATA IN EXCEL FILE
import selenium
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import undetected_chromedriver as uc
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd
import time 
import re
import csv
import time
import os
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
import random



# Set up Chrome options
chrome_options = Options()
#chrome_options.add_argument("--headless")  # Run in headless mode
chrome_options.add_argument("--disable-blink-features=AutomationControlled")
chrome_options.add_argument("--disable-infobars")
chrome_options.add_argument("--incognito")
chrome_options.add_argument("--start-maximized")
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--disable-dev-shm-usage")  # Overcome limited resource problems
chrome_options.add_argument("--window-size=1920x1080")
sleep=time.sleep
# Initialize the WebDriver
drivar = uc.Chrome(options=chrome_options)
sleep=time.sleep
actions = ActionChains(drivar)
actions.move_by_offset(50, 50).perform()  # Random mouse movement
time.sleep(random.uniform(1, 3))
# Navigate to the URL
url = 'https://de.indeed.com/'
drivar.get(url)
sleep(15)
print(drivar.title)
countery='Германия'
wait=WebDriverWait(drivar,10)
def captcha(drivar):
    try:
        while True:
            # Wait for the page to load or some specific element to appear
            WebDriverWait(drivar, 10).until(EC.presence_of_element_located((By.XPATH, '//body')))
            # Check if captcha is present on the page
            if 'Additional Verification Required' in drivar.page_source:
                print("Captcha detected. Attempting to solve...")
                # Wait until the captcha is clickable, then click the captcha or necessary element
                captcha_element = WebDriverWait(drivar, 10).until(EC.element_to_be_clickable((By.XPATH, '//div[@class="main-content"]')))
                captcha_element.click()
                print("Captcha solved successfully.")
                time.sleep(5)  # Shorter sleep or replace with an appropriate wait
            else:
                print("No captcha detected. Exiting loop.")
                break  # Exit the loop if captcha is not detected

    except:
        print(f"An unexpected error occurred")


if "Additional Verification Required" in drivar.page_source:                      
   captcha()
   sleep(5)
else:
     print("Captcha Not accer --- ")

try:
    while True:
        
        if "Alle ablehnen" in drivar.page_source:
                wait.until(EC.element_to_be_clickable((By.ID,'onetrust-reject-all-handler'))).click()    # Reject Cookes
                print("Cookes Accept")
        else:
            print("Cookes Not Found")
            break
except:
        print("Cookes Not Found")
            
sleep(2) 
location=wait.until(EC.element_to_be_clickable((By.ID,'text-input-where')))
location.send_keys(countery)     # Countery Name
sleep(1)     
job_titles = pd.read_csv('data.csv')['Keywords'].tolist()

#---------------------------------------------------------------------------------------------------
job_titles_new = job_titles[4:6]  # This will take titles at index 1 and 2      ((5 to 8))         |
#---------------------------------------------------------------------------------------------------

data = {
    'Job Links': []  
}
unique_job_links=set()

wait=WebDriverWait(drivar,7)

for job_title in job_titles_new:
    # Clear the previous input and enter new job title
    try:
      captcha()
    except:
        # Maximize the browser window
        drivar.maximize_window()
        print("Ensuring the Browser Window is Maximized and Active ")
        sleep(1)
        input_field=wait.until(EC.element_to_be_clickable((By.ID, 'text-input-what')))
        input_field.click()
        sleep(1)
        input_field.send_keys(Keys.CONTROL + 'a')  # Select all text
        sleep(1)
        input_field.send_keys(Keys.DELETE)          # Delete selected text
        sleep(1)
        wait.until(EC.element_to_be_clickable((By.ID, 'text-input-what'))).send_keys(job_title, Keys.ENTER)
        print("At This Time Process Job Tital: ", job_title)
        sleep(1)  # Wait for the page to load
    try:
            # Move Hear
        while True:
                # Wait for the main job results div to be clickabl
                
                par = wait.until(EC.presence_of_element_located((By.XPATH,'//*[@id="mosaic-provider-jobcards"]/ul')))
                all_li = par.find_elements(By.TAG_NAME,'h2')
                for i in all_li:
                    link  = i.find_element(By.TAG_NAME,'a').get_attribute("href") 
                    print("URL Not Present in File, Unique Link Retrieved Successfully: ", link)
                    unique_job_links.add(link)
                    #data['Job Links'].append(link)
                        
                # Now check for the next page only after collecting data
                try:
                    # sleep(3)
                    element=wait.until(EC.presence_of_element_located((By.XPATH,'//a[@aria-label="Next Page"]')))
                    drivar.execute_script("arguments[0].scrollIntoView(true);", element)
                    # Ensure the element is visible and enabled before clicking
                    if element.is_displayed() and element.is_enabled():
                        drivar.execute_script("arguments[0].click();", element)
                    else:
                        print("Element is not clickable.")

                    sleep(1)
                    print("Move To Next Page Successfully ----- ")

                except:
                        if "Additional Verification Required" in drivar.page_source:
                          captcha()
                        else:
                            print("No More Page Found In This URL")
                            break
    except:
         print("Process intrapted for few seconds, Sorry!  ")                

# Optionally print or save the collected links
print("All linkes Save SuccessFully")
wait=WebDriverWait(drivar,5)


# Function to create a new Excel file with headers if it doesn't exist
def create_excel_file(file_path):
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Job Data"
    headers = ['Company Name', 'Website Link', 'Revenue', "JOB TITLE", "JOB LINK", "E-MAIL", "PHONE"]
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        sheet[f'{col_letter}1'] = header
    wb.save(file_path)

# Function to write data to an existing Excel file, appending in the next row
def data_send_to_excel_file(file_path, company_n, website_url, rav, Job_tital, job_link, email_id, phone_no):
    try:
        # Try to load the existing workbook
        wb = load_workbook(file_path)
    except FileNotFoundError:
        # If file doesn't exist, create a new workbook and add headers
        create_excel_file(file_path)
        wb = load_workbook(file_path)

    sheet = wb.active
    
    # Check if headers exist (look for the first non-empty row)
    if sheet.max_row == 1 and sheet.cell(1, 1).value is None:
        # No headers found, so we need to add headers
        headers = ['Company Name', 'Website Link', 'Revenue', "JOB TITLE", "JOB LINK", "E-MAIL", "PHONE"]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            sheet[f'{col_letter}1'] = header
    
    # Append data to the next row
    next_row = sheet.max_row + 1
    sheet[f'A{next_row}'] = company_n
    sheet[f'B{next_row}'] = website_url
    sheet[f'C{next_row}'] = rav
    sheet[f'D{next_row}'] = Job_tital
    sheet[f'E{next_row}'] = job_link
    sheet[f'F{next_row}'] = email_id
    sheet[f'G{next_row}'] = phone_no

    wb.save(file_path)
    print(f"Data saved successfully to {file_path}")
    # Path to your Excel file
    file_path = 'result.xlsx'
        # Open each job link one by one
#for job_link in data['Job Links']:
total_search_jobs=0
unique_job_tiitals_in_excal_file=0
for job_link in unique_job_links:
    try:

        total_search_jobs+=1
        phone_no=''
        email_id=''
        rav=''
        website_url=''
        company_n=''
        Job_tital=''

        drivar.get(job_link)
        sleep(2)  # Wait for the job page to load
        print("Total Searched Job's at this time: ", total_search_jobs)
        
        try:
            try:
                try:
                    print("Try to get Job Tital")
                    sleep(1)
                    h1 = wait.until(EC.element_to_be_clickable((By.XPATH, '//h1[@class="jobsearch-JobInfoHeader-title css-10fqp5z e1tiznh50"]')))
                    span = h1.find_element(By.TAG_NAME, 'span')
                    Job_tital1 = span.text
                    print("Job Tital :",Job_tital1)
                    print("Try To Read Excel File")
                except:
                        sleep(2)
                        if "Additional Verification Required" in drivar.page_source:   
                                    captcha()                   
                        else: 

                            print("Refresh and Try again to get Job Tital")
                            drivar.refresh()
                            sleep(2)
                            h1 = wait.until(EC.element_to_be_clickable((By.XPATH, '//h1[@class="jobsearch-JobInfoHeader-title css-10fqp5z e1tiznh50"]')))
                            span = h1.find_element(By.TAG_NAME, 'span')
                            Job_tital1 = span.text
                            print("Job Tital :",Job_tital1)
                            print("Try To Read Excel File")

                def read_job_data(file_path):
                    job_data = []  # List to store job titles
                    try:
                        # Load the workbook and sheet
                        wb = openpyxl.load_workbook(file_path)
                        sheet = wb.active  # Get the active sheet (or specify by name if needed)
                        
                        # Loop through rows starting from the second row (skipping the header)
                        for row in sheet.iter_rows(min_row=2, min_col=4, max_col=4, values_only=True):  # min_col=4 and max_col=4 to get the 4th column (index 3)
                            job_data.append(row[0])  # row[0] will hold the value from the 4th column (Job Title)
                        wb.close()
                    except Exception as e:
                        print(f"Error reading the Excel file: {e}")
                    
                    return job_data

                # File path to the Excel file
                excel_file_path = 'result.xlsx'

                # Call the function to read the job data
                job_data = read_job_data(excel_file_path)

                # Check if the specific job title already exists
                if Job_tital1 in job_data:
                    print("Job Title already present in file")
                    print("--- So move to other link ---")
                else:
                    unique_job_tiitals_in_excal_file +=1
                    print("Unique title of job so save this:", Job_tital1)  # Job title
                    print("Unique Founded Jobs Data So Far :", unique_job_tiitals_in_excal_file)
                    Job_tital = Job_tital1

                    # ----------------------------------------------------------------------------
                        # Find Phone Nmber
                    try:
                        page_text = drivar.page_source  # Correct the spelling of 'driver' from 'drivar'
                        sleep(1)
                        # Define regex patterns for phone numbers with more than 8 digits
                        phone_pattern = r'\+?\d{1,3}\s?\d{1,5}[\s-]?(\d{2,4}[\s-]?){2,}'
                        email_pattern = r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'

                        # Find all phone numbers
                        phone_numbers = re.findall(phone_pattern, page_text)

                        # Filter phone numbers to include only those with more than 8 digits
                        valid_phone_numbers = [
                            number for number in phone_numbers 
                            if len(re.sub(r'\D', '', number)) > 8
                        ]

                        # Prefer phone numbers that start with a '+' sign
                        preferred_numbers = [number for number in valid_phone_numbers if number.startswith('+')]

                        # Select the last valid phone number, preferring those starting with '+'
                        if preferred_numbers:
                            selected_phone_number = preferred_numbers[0]  # first preferred number
                        else:
                            selected_phone_number = valid_phone_numbers[0] if valid_phone_numbers else None

                        # Find all email addresses
                        email_addresses = re.findall(email_pattern, page_text)

                        # Select the last email address
                        selected_email_address = email_addresses[0] if email_addresses else None
                        sleep(2)
                        # Print results
                        if selected_phone_number:
                                
                            print("Phone Number:", selected_phone_number)
                            data['Phone No'].append(selected_phone_number)
                            phone_no=selected_phone_number
                        else:
                            print("Phone Number:", "Not Found")
                            # data['Phone No'].append("Phone Not Found")    

                        if selected_email_address:
                                if selected_email_address == 'noreply@indeed.com':
                                    print("E-mail address:", "Not Found")
                                    # data['E-Mail'].append("E-mail Not Found")
                                else:
                                    print("Founded E-Mail Address:", selected_email_address)
                                    email_id=selected_email_address
                        else:
                            print("E-mail Number:", "Not Found")
                            # data['E-Mail'].append("E-mail Not Found")

                    except:
                            print("Phone Number:", "Not Found")
                            # data['Phone No'].append("Phone Not Found")
                            print("E-mail Number:", "Not Found")
                            # data['E-Mail'].append("E-mail Not Found") 
    

                    try:
                        # click for find Website link and Company Name
                        print("Try To click for Move to Next Page ----  ")
                        try:                     
                            click_to_find=wait.until(EC.presence_of_element_located((By.XPATH,'//a[@class="css-1gcjz36 e19afand0"]')))
                            click_to_find.click()
                            sleep(2)
                            # Get current window handles
                            handles = drivar.window_handles
                            # Switch to the new tab
                            drivar.switch_to.window(handles[1])
                        except:
                            drivar.refresh()
                            print("Again Try to Click -----")
                            sleep(2)
                            click_to_find=wait.until(EC.presence_of_element_located((By.XPATH,'//a[@class="css-1gcjz36 e19afand0"]')))
                            click_to_find.click()
                            sleep(2)
                            # Get current window handles
                            handles = drivar.window_handles
                            # Switch to the new tab
                            drivar.switch_to.window(handles[1])
                            # Get Company Name
                        try:
                            sleep(1)
                            company_name_hader=wait.until(EC.element_to_be_clickable((By.XPATH,'//header[@class="css-1vcnma3 eu4oa1w0"]')))
                            company_name=company_name_hader.find_element(By.XPATH,'//div[@class="css-1wvasm5 e1wnkr790"]').text
                            print("company name :", company_name)                 # Name of Company
                            company_n=company_name
                        except:
                                print("company name :","Not found")
                                # data['Company Name'].append("company name Not Found")

                        # get Company Website Link
                        try:
                            wait=WebDriverWait(drivar,2)
                            parent_for_get_company_link=wait.until(EC.element_to_be_clickable((By.XPATH,'//section[@class="css-dg0oyg eu4oa1w0"]')))
                            a_div_in_this_parent=parent_for_get_company_link.find_element(By.XPATH,'//div[@class="css-1r0tpua eu4oa1w0"]')
                            ul=a_div_in_this_parent.find_element(By.XPATH,'//ul[@class="css-hbpv4x e37uo190"]')
                            li=ul.find_element(By.XPATH,'//li[@data-testid="companyInfo-companyWebsite"]')
                            in_this_li_a_div=li.find_element(By.XPATH,'//div[@class="css-kaq73 e37uo190"]')
                            # Now find the anchor tag within this div
                            anchor_tag = in_this_li_a_div.find_element(By.TAG_NAME, 'a')
                            # Get the href attribute
                            website_link = anchor_tag.get_attribute('href')
                            print("Websitr Link: ",website_link) 
                            website_url=website_link
                            # data['Website Link'].append(website_link) 
                        # WEBSITE LINK
                        except:
                                print("Websitr Link: ","Not Found")
                                
                        try:

                                    print("Try to found revenue from NorthData.")
                                    drivar.get('https://www.northdata.de/_login?targetUrl=https%3A%2F%2Fwww.northdata.de%2F')
                                    print("   -------  Go to Northh Website  ---------")
                                    sleep=time.sleep
                                    sleep(5)
                                    wait=WebDriverWait(drivar,10)
                                    email=wait.until(EC.presence_of_element_located((By.XPATH,'//input[@name="email"]' )))
                                    sleep(1)
                                    email.click()
                                    sleep(2)
                                    email.send_keys('j.krolova@hekler.ee')
                                    password=wait.until(EC.presence_of_element_located((By.XPATH, '//input[@name="password"]')))
                                    password.click()
                                    sleep(2)
                                    password.send_keys('HKRManpower2024!')
                                    sleep(1)
                                    wait.until(EC.presence_of_element_located((By.XPATH,'//button[@type="submit"]'))).click()
                                    print("Tital: ",drivar.title)
                                    wait = WebDriverWait(drivar, 5)
                                    sleep(3)  # Allow some time for the page to load
                                    try:
                                        wait.until(EC.presence_of_element_located((By.XPATH,'//span[@id="cmpbntyestxt"]'))).click()
                                    except:
                                        print("Cookes Not Accer")

                                    sleep(1)

                                    wait = WebDriverWait(drivar, 10)  

                                    # Locate the input field for company search
                                    input_fild = wait.until(EC.presence_of_element_located((By.XPATH, '//input[@type="search"]')))
                                    input_fild.click()
                                    sleep(1)
                                    # Type in the company name
                                    input_fild.send_keys(company_name)
                                    input_fild.send_keys(Keys.ENTER)
                                    sleep(4)
                                    umt=wait.until(EC.presence_of_element_located((By.XPATH,'//a[@data-tab="tab-bc-1"]')))
                                    drivar.execute_script("arguments[0].scrollIntoView();", umt)
                                    sleep(2)
                                    print(" ------ Click On UMSATZ ------- ")
                                    umt.click()
                                    sleep(3)
                                    main_div=wait.until(EC.presence_of_element_located((By.XPATH, '//div[@data-item="Revenue"]')))
                                    sleep(1)
                                    last_rect_element = drivar.execute_script("""

                                        let mainDiv = arguments[0];
                                        let rects = mainDiv.querySelectorAll('svg rect');
                                        if (rects.length > 0) {
                                            return rects[rects.length - 1];  // Return last rect element
                                        }

                                        return null;  // Return null if no rect elements

                                    """, main_div)

                                    if last_rect_element:

                                        print("Last rect element found:")
                                        drivar.execute_script("arguments[0].scrollIntoView();", last_rect_element)
                                        sleep(1)
                                        last_rect_element.click()
                                        print("Clicked on the rect element ---  click --------- click -----.")
                                        sleep(3)
                                        x=drivar.find_element(By.XPATH,'//div[@class="tooltip ui basic black label right pointing"]')
                                        a=x.find_element(By.XPATH,'//span[@class="value"]')
                                        rav=a.text
                                        print("Get Ravinew From Groph: ", a.text)
                                    else:
                                        print("No valid rect element found.")
                                        rav="Ravinew Not Found"
                                    page_source = drivar.page_source

                                    try:
                                        # Regular Expression to match phone numbers like +49 895480340
                                        phone_number_regex = r'\+49\s?\d{7,14}'
                                        # Find all phone numbers matching the pattern
                                        phone_numbers = re.findall(phone_number_regex, page_source)
                                        number=phone_numbers[0]
                                        print("Founded Phone:", number)
                                        phone_no=number

                                    except:

                                         print("Phone Number Not Found")
                                    # e-mail
                                    try:
                                        # Regular Expression to match phone numbers like +49 895480340
                                        email_pattern = r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
                                        # Find all phone numbers matching the pattern
                                        emails = re.findall(email_pattern, page_source)
                                        for email in emails:
                                            if 'j.krolova@hekler.ee' == email:
                                                pass
                                            else:
                                                print("Unique E-MAil Found:  ",email)
                                                email_id=email
                                                break
                                    except:
                                         print("E-Mail Not Found ")

                        except :
                                    print("Error while retrieving revenue: ")
                        # Closing the tab and switching back to the original window after the process
                        finally:
                                try:
                                    drivar.close()
                                    handles = drivar.window_handles
                                    drivar.switch_to.window(handles[0])  # Switch back to the original tab
                                except Exception as e:

                                    print("Error closing or switching tabs: ", e)   

                    except:
                        print("Not Move to Next Page")
                    file_path = 'result.xlsx'
                    try:
                        file_path = 'result.xlsx'
                        data_send_to_excel_file(file_path, company_n, website_url, rav, Job_tital, job_link, email_id, phone_no)
                        print("Data Saved into Excel File ---")
                        print("Total Job's Details So Far:", unique_job_tiitals_in_excal_file)
                    except Exception as e:
                        print(f"Data Not Saved --- Error: {e}")

                        #-----------------------------------------------------------------------------
            except:
                        try:
                            if "Additional Verification Required" in drivar.page_source:   
                                    captcha() 
                        except:
                            print("Internat Issue ---------- ")
                                    
        except Exception as e:
                print(f"Could not perform action on {job_link}: {e}")
    except:
            try:
                if "Additional Verification Required" in drivar.page_source:   
                                    captcha() 
            except:
                 print("Internat Issue ---- ")

            # Get all iframes on the page

print("Job data saved to data.csv successfully.")
# Close the driver
drivar.quit()