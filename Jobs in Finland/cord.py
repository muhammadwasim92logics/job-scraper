# SAvE DATA INTO EXCEL FILE 
import selenium
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import undetected_chromedriver as uc
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd
import time
import re
import csv
from selenium.webdriver.chrome.options import Options
import openpyxl
from bs4 import BeautifulSoup


# Set up Chrome options
chrome_options = Options()
chrome_options.add_argument("--incognito")
chrome_options.add_argument("--disable-dev-shm-usage")
chrome_options.add_argument("--disable-blink-features=AutomationControlled")
chrome_options.add_argument("--disable-infobars")
chrome_options.add_argument("--start-maximized")
chrome_options.add_argument("--window-size=1920x1080")
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--disable-dev-shm-usage")  # Overcome limited resource problems
chrome_options.add_argument("--window-size=1920x1080")  # Set window size (optional)
# chrome_options.add_argument("--headless")

# Initialize the WebDriver
drivar = uc.Chrome(options=chrome_options)

url = 'http://www.duunitori.fi'
drivar.get(url)

sleep = time.sleep
countery = 'Finland'
wait = WebDriverWait(drivar, 6)
print(drivar.title)
sleep(3)
wait.until(EC.element_to_be_clickable((By.XPATH, '//button[@class="btn btn--primary gdpr-modal__button gdpr-modal__button--accept gdpr-close"]'))).click()  # Reject Cookies
sleep(2)
main = wait.until(EC.element_to_be_clickable((By.XPATH, '(//input[@class="taggle_input"])[2]')))
main.clear()
sleep(2)
main.send_keys(countery)  # Country Name
sleep(2)

# Read job titles from CSV file
job_titles = pd.read_csv('data.csv')['Keywords'].tolist()


#----------------------------------------------------------------------------------------------------------
job_titles_new = job_titles[0:5]  # This will take titles at index 1 and 2                                |
#----------------------------------------------------------------------------------------------------------



# Data structure to hold unique job links
unique_links = set()

# Loop through job titles and search
for job_title in job_titles_new:
    sleep(2)
    input_field = wait.until(EC.element_to_be_clickable((By.XPATH, '//input[@class="taggle_input"]')))
    sleep(1)
    input_field.send_keys(Keys.BACKSPACE, Keys.BACKSPACE)
    sleep(1)
    input_field.send_keys(Keys.BACKSPACE, Keys.BACKSPACE)
    sleep(1)
    input_field.send_keys(job_title, Keys.ENTER)  # keyword
    print("At This Time Process Job Title: ", job_title)
    sleep(2)  # Wait for the page to load

    try:
        while True:
            sleep(2)
            print("Try to Get Job Links")
            par = wait.until(EC.presence_of_element_located((By.XPATH, '//div[@class="grid-sandbox grid-sandbox--tight-bottom grid-sandbox--tight-top"]')))
            all_li = par.find_element(By.XPATH, '(//div[@class="grid"])[4]')
            divs = all_li.find_elements(By.XPATH, ".//div[normalize-space(@class) = 'grid grid--middle job-box job-box--lg']")

            for open_one_by_one in divs:
                # Use a relative XPath to find the <a> element within the current job box
                link = open_one_by_one.find_element(By.XPATH, './/a[@class="job-box__hover gtm-search-result"]').get_attribute('href')
                unique_links.add(link)  # Add link to the set
                print("Job link:", link)
            try:
                # Wait until the button is present and clickable
                button_without_title = WebDriverWait(drivar, 8).until(EC.element_to_be_clickable((By.XPATH, '//a[@class="pagination__page-round" and not(@title)]')))
                drivar.execute_script("arguments[0].click();", button_without_title)
                print("Clicked")
            except Exception as e:
                print("No Move Page Available")
                break
            finally:
                print("Execution complete")
    except:
        print("Not Process This Keyword")

# Function to create a new Excel file with headers if it doesn't exist
def create_excel_file(file_path):
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Job Data"
    headers = ['Company Name', 'Website Link', 'Revenue', "JOB TITLE", "JOB LINK", "E-MAIL", "PHONE"]
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        sheet[f'{col_letter}1'] = header
    wb.save(file_path)

# Function to write data to an existing Excel file, appending in the next row
def data_send_to_excel_file(file_path, company_n, website_url, rav, Job_tital, job_link, email_id, phone_no):
    try:
        # Try to load the existing workbook
        wb = load_workbook(file_path)
    except FileNotFoundError:
        # If file doesn't exist, create a new workbook and add headers
        create_excel_file(file_path)
        wb = load_workbook(file_path)

    sheet = wb.active
    
    # Check if headers exist (look for the first non-empty row)
    if sheet.max_row == 1 and sheet.cell(1, 1).value is None:
        # No headers found, so we need to add headers
        headers = ['Company Name', 'Website Link', 'Revenue', "JOB TITLE", "JOB LINK", "E-MAIL", "PHONE"]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            sheet[f'{col_letter}1'] = header
    
    # Append data to the next row
    next_row = sheet.max_row + 1
    sheet[f'A{next_row}'] = company_n
    sheet[f'B{next_row}'] = website_url
    sheet[f'C{next_row}'] = rav
    sheet[f'D{next_row}'] = Job_tital
    sheet[f'E{next_row}'] = job_link
    sheet[f'F{next_row}'] = email_id
    sheet[f'G{next_row}'] = phone_no

    wb.save(file_path)
    print(f"Data saved successfully to {file_path}")
try:
    # Path to your Excel file
    file_path = 'result.xlsx'

    # Open each job link one by one
    for job_link in unique_links:
        phone_no = ''
        email_id = ''
        rav = ''
        website_url = ''
        company_n = ''
        Job_tital = ''
        post_date = ''
        joblink = ''

        drivar.get(job_link)
        sleep(2)  # Wait for the job page to load
        
        try:
            try:
                sleep(1)
                print("Try to get Job Title")
                h1 = wait.until(EC.presence_of_element_located((By.XPATH, '//h1[@class="text--break-word"]'))).text
                Job_tital1 = h1
                print("Job Title: ", h1)  # Job title
            except:
                print("Refresh And Try Again")
                drivar.refresh()
                sleep(1)
                print("Try to get Job Title")
                h1 = wait.until(EC.presence_of_element_located((By.XPATH, '//h1[@class="text--break-word"]'))).text
                Job_tital1 = h1
                print("Job Title: ", h1)  # Job title

            print("Try To Read Excel File")
            def read_job_data(file_path):
                job_data = []  # List to store job titles
                try:
                    # Load the workbook and sheet
                    wb = openpyxl.load_workbook(file_path)
                    sheet = wb.active  # Get the active sheet (or specify by name if needed)
                    
                    # Loop through rows starting from the second row (skipping the header)
                    for row in sheet.iter_rows(min_row=2, min_col=4, max_col=4, values_only=True):  # min_col=4 and max_col=4 to get the 4th column (index 3)
                        job_data.append(row[0])  # row[0] will hold the value from the 4th column (Job Title)
                    wb.close()
                except Exception as e:
                    print(f"Error reading the Excel file: {e}")
                
                return job_data

            # File path to the Excel file
            excel_file_path = 'result.xlsx'

            # Call the function to read the job data
            job_data = read_job_data(excel_file_path)

            # Check if the specific job title already exists
            if Job_tital1 in job_data:
                print("Job Title already present in file")
                print("--- So move to other link ---")
            else:
                print("Unique title of job so save this:", Job_tital1)  # Job title
                Job_tital = Job_tital1

                
                # Phone and E-Mail
                page_text = drivar.page_source
                phone_pattern = r'\d{3,4} \d{3,4} \d{3,4}'
                email_pattern = r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'

                try:
                    email_address = re.search(email_pattern, page_text)
                    first_email_address = email_address.group()
                    print("Email:", first_email_address)
                    # email_id = first_email_address
                    if email_id == 'noreply@indeed.com':
                        pass
                    else:
                        email_id = first_email_address
                        
                except:
                    print("Error finding email address:")

                try:
                    phone_no = re.search(phone_pattern, page_text)
                    first_phone_no = phone_no.group()
                    print("Phone:", first_phone_no)
                    phone_no = first_phone_no
                except:
                    print("Error finding phone number:")

                try:
                    P_date = drivar.find_element(By.XPATH, '(//p[@class="header__info"])[2]').text
                    post_date = P_date
                    print("Job Posting Date:", post_date)
                except:
                    print("Posting Date Not Found")

                try:
                    print("Try To Found Company Name")
                    company_name_main = drivar.find_element(By.XPATH, '//p[@class="header__info"]')
                    span_elements = company_name_main.find_elements(By.TAG_NAME, 'span')

                    # Loop through the <span> tags and check for text
                    text_found = False
                    for span_element in span_elements:
                        span_text = span_element.text.strip()  # Get the text and remove any extra spaces
                        
                        # If text is found, print it and break the loop
                        if span_text:
                            print("Company Name Found:", span_text)
                            company_n = span_text
                            text_found = True
                            break
                except:
                    print("Company Name Not Found")


                # Try to find Ravinew
                try:
                    wait1=WebDriverWait(drivar,1)
                    ravinew = wait1.until(EC.presence_of_element_located((By.XPATH, '//span[@class="financial__value"]'))).text
                    print("Revenue:", ravinew)
                    rav = ravinew
                except:

                    print("Ravinew Not Found, so Go to Ravinew Finder Website And Find It")
                    url1 = f'https://www.finder.fi/search?what={span_text}'
                    print("Company Revenue: ", span_text)
                    drivar.get(url1)
                    sleep(1)
                    drivar.refresh()

                    try:
                        sleep(1)
                        revenue_element = wait.until(EC.presence_of_element_located((By.XPATH, '//h3[@class="MuiTypography-root MuiTypography-h6 css-1wn89e2"]'))).text
                        print("Revenue Found:", revenue_element)
                        rav = revenue_element
                    except:
                        rav = 'Revenue not found'
                        print("Revenue also not found at Revenue Finder site")



                #------------------------------------------------------------------------------------------
                #  Find Website URL
                try: 
                        drivar.get(job_link)
                        sleep(2)
                        page_source=drivar.page_source
                        soup = BeautifulSoup(page_source, 'html.parser')
                        visible_text = soup.get_text()
                        domain_pattern = r'\bwww\.[a-zA-Z0-9-]+\.[a-zA-Z]{2,}\b'
                        web_url = re.findall(domain_pattern, visible_text)
                        print("Try to found Website URL for Main Page")
                        if web_url:
                            print("Website URL Found: ", web_url)
                            print("Checking URL --- ")
                            #website_url = web_url[0]  # Assuming the first match is what you want
                            Founded_website_url = web_url[0]
                            if Founded_website_url == 'www.linkedin.com':
                                print("URL is www.Linkedin.com")
                                print("So move to next page and find Website URL")

                                try:
                                    # Move Next Page For fiding website Link
                                        print("Eebsite URL Not Found In Main Page, Try To click for Move to Next Page ----  ")
                                        moving=drivar.find_element(By.XPATH,'//div[@class="1/1 grid__cell btn-group btn-group--center"]')
                                        move=moving.find_element(By.TAG_NAME,'a')
                                        drivar.execute_script("arguments[0].click();", move)
                                        print("Clicked and move to next Page")
                                        sleep(2)

                                                        # Website Link
                                        try:    
                                                wait2=WebDriverWait(drivar,2)
                                                website_link_main=wait2.until(EC.presence_of_element_located((By.XPATH,'//div[@class="  pcp__employer-info__buttons-div  "]')))
                                                website_link=website_link_main.find_element(By.TAG_NAME,'a').get_attribute('href')
                                                print("Website URL 1 Place :",website_link)
                                                website_url=website_link
                                        except:
                                                wait3=WebDriverWait(drivar,1)
                                                website_link2= drivar.find_element(By.XPATH,'//div[@itemprop="url"]')
                                                website3=website_link2.find_element(By.TAG_NAME,'strong')
                                                website_2nd=website3.find_element(By.TAG_NAME,'a').get_attribute('href')
                                                print("website URL 2 Place:",website_2nd)                    
                                                website_url=website_2nd
                                except:
                                            print("Next Page Not Avilable ")
                            else:
                                website_url = web_url[0]
                        else:
                                try:
                                    # Move Next Page For fiding website Link
                                        print("Eebsite URL Not Found In Main Page, Try To click for Move to Next Page ----  ")
                                        moving=drivar.find_element(By.XPATH,'//div[@class="1/1 grid__cell btn-group btn-group--center"]')
                                        move=moving.find_element(By.TAG_NAME,'a')
                                        drivar.execute_script("arguments[0].click();", move)
                                        print("Clicked and move to next Page")
                                        sleep(2)

                                                        # Website Link
                                        try:    
                                                wait2=WebDriverWait(drivar,2)
                                                website_link_main=wait2.until(EC.presence_of_element_located((By.XPATH,'//div[@class="  pcp__employer-info__buttons-div  "]')))
                                                website_link=website_link_main.find_element(By.TAG_NAME,'a').get_attribute('href')
                                                print("Website URL 1 Place :",website_link)
                                                website_url=website_link
                                        except:
                                                wait3=WebDriverWait(drivar,1)
                                                website_link2= drivar.find_element(By.XPATH,'//div[@itemprop="url"]')
                                                website3=website_link2.find_element(By.TAG_NAME,'strong')
                                                website_2nd=website3.find_element(By.TAG_NAME,'a').get_attribute('href')
                                                print("website URL 2 Place:",website_2nd)                    
                                                website_url=website_2nd
                                except:
                                            print("Next Page Not Avilable ")
                except:
                        print("Error In finding Website URL")


    # ------------------------------------------------------------------------------------------------
                # try:
                #     wait1=WebDriverWait(drivar,1)
                #     ravinew = wait1.until(EC.presence_of_element_located((By.XPATH, '//span[@class="financial__value"]'))).text
                #     print("Revenue:", ravinew)
                #     rav = ravinew
                # except:

                #     print("Ravinew Not Found, so Go to Ravinew Finder Website And Find It")
                #     url1 = f'https://www.finder.fi/search?what={span_text}'
                #     print("Company Revenue: ", span_text)
                #     drivar.get(url1)
                #     sleep(1)
                #     drivar.refresh()

                #     try:
                #         sleep(1)
                #         revenue_element = wait.until(EC.presence_of_element_located((By.XPATH, '//h3[@class="MuiTypography-root MuiTypography-h6 css-1wn89e2"]'))).text
                #         print("Revenue Found:", revenue_element)
                #         rav = revenue_element
                #     except:
                #         rav = 'Revenue not found'
                #         print("Revenue also not found at Revenue Finder site")
                # try:
                #         drivar.get(job_link)
                #         time.sleep(1)
                #         page_source=drivar.page_source
                #         soup = BeautifulSoup(page_source, 'html.parser')
                #         visible_text = soup.get_text()
                #         domain_pattern = r'\bwww\.[a-zA-Z0-9-]+\.[a-zA-Z]{2,}\b'
                #         web_url = re.findall(domain_pattern, visible_text)
                #         print("Try to found Website URL for Main Page")
                #         if web_url:
                #            not_conceder_url=['www.linkedin.com', 'www.facebook.com', 'www.instagram.com']
                #            if web_url in not_conceder_url:
                #                 print("Not Proper Url Founde Founded URL are (linkedin,facebook,Instagram) So Not Save These URLS")
                #            else:
                #             print("Website URL Found: ", web_url)
                #             print("No Need TO Move NExt Page Now")
                #             website_url=web_url
                #         else:
                #                 try:
                #                     # Move Next Page For fiding website Link
                #                         print("Eebsite URL Not Found In Main Page, Try To click for Move to Next Page ----  ")
                #                         moving=wait.until(EC.presence_of_element_located((By.XPATH,'//div[@class="1/1 grid__cell btn-group btn-group--center"]')))
                #                         move=moving.find_element(By.TAG_NAME,'a')
                #                         drivar.execute_script("arguments[0].click();", move)
                #                         print("Clicked and move to next Page")
                #                         sleep(2)

                #                                         # Website Link
                #                         try:    
                #                                 wait=WebDriverWait(drivar,3)
                #                                 website_link_main=wait.until(EC.presence_of_element_located((By.XPATH,'//div[@class="  pcp__employer-info__buttons-div  "]')))
                #                                 website_link=website_link_main.find_element(By.TAG_NAME,'a').get_attribute('href')
                #                                 print("Website URL 1 Place :",website_link)
                #                                 website_url=website_link
                #                         except:
                #                                 wait=WebDriverWait(drivar,3)
                #                                 website_link2=wait.until(EC.presence_of_element_located((By.XPATH,'//div[@itemprop="url"]')))
                #                                 website3=website_link2.find_element(By.TAG_NAME,'strong')
                #                                 website_2nd=website3.find_element(By.TAG_NAME,'a').get_attribute('href')
                #                                 print("website URL 2 Place:",website_2nd)                    
                #                                 website_url=website_2nd
                #                 except:
                #                             print("Next Page Not Avilable ")
                # except:
                #         print("Error In finding Website URL")

                try:
                    data_send_to_excel_file(file_path, company_n, website_url, rav, Job_tital, job_link, email_id, phone_no)
                    print("Data Saved ---")
                except:
                    print("Data Not Saved ---")
        except Exception as e:
            print(f"Could not perform action on {job_link}: {e}")

    print("Job data saved to result.xlsx successfully.")
except:
     print("Connection Issue, Move to Next")
# Close the driver
drivar.quit()
